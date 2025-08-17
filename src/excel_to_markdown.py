import os
from pathlib import Path
from markitdown import MarkItDown
import openpyxl # openpyxlをインポート

def convert_excel_to_markdown():
    """
    doc/excel フォルダ内のすべての Excel ファイルを検索し、
    Markdown に変換して doc/markdown フォルダに保存する。
    1シート＝1Markdownファイルとして出力する。
    """
    excel_dir = Path("doc/excel")
    markdown_dir = Path("doc/markdown")

    # 出力先フォルダが存在しない場合は作成
    markdown_dir.mkdir(exist_ok=True)

    # markitdownのインスタンスを作成 (今回は直接使用しないが、将来的な拡張のために残す)
    # md_converter = MarkItDown(enable_plugins=False)

    # 許可する拡張子
    allowed_extensions = [".xlsx", ".xls", ".xlsm"]

    print(f"'{excel_dir}' 内のExcelファイルを検索しています...")

    # excel_dir 内のファイルを再帰的に検索
    found_files = False
    for excel_file in excel_dir.glob("**/*"):
        if excel_file.is_file() and excel_file.suffix.lower() in allowed_extensions:
            found_files = True
            print(f"変換中: {excel_file}")

            try:
                # openpyxlでワークブックを開く
                workbook = openpyxl.load_workbook(excel_file)

                # 各シートを処理
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    markdown_content = ""

                    # シートの内容をMarkdownテーブルとして整形
                    # ヘッダー行の取得
                    header = [cell.value if cell.value is not None else "" for cell in sheet[1]]
                    if header:
                        markdown_content += "| " + " | ".join(map(str, header)) + " |\n"
                        markdown_content += "| " + " | ".join(["---"] * len(header)) + " |\n"

                    # データ行の取得
                    for row_index, row in enumerate(sheet.iter_rows(min_row=2)): # ヘッダー行の次から開始
                        row_values = [cell.value if cell.value is not None else "" for cell in row]
                        # 空行をスキップ
                        if all(value == "" for value in row_values):
                            continue
                        markdown_content += "| " + " | ".join(map(str, row_values)) + " |\n"

                    # 出力ファイルパスを決定
                    # 元のディレクトリ構造を維持し、ファイル名にシート名を追加
                    relative_path = excel_file.relative_to(excel_dir)
                    # 拡張子を除いたファイル名を取得
                    base_name = relative_path.stem
                    # シート名をファイル名に追加
                    markdown_file_name = f"{base_name}_{sheet_name}.md"
                    markdown_file_path = markdown_dir / relative_path.parent / markdown_file_name

                    # 出力先のサブディレクトリを作成
                    markdown_file_path.parent.mkdir(parents=True, exist_ok=True)

                    # Markdown ファイルを書き込み
                    with open(markdown_file_path, "w", encoding="utf-8") as f:
                        f.write(markdown_content)

                    print(f"保存しました: {markdown_file_path}")

            except Exception as e:
                print(f"エラー: {excel_file} の変換に失敗しました - {e}")

    if not found_files:
        print("変換対象のExcelファイルが見つかりませんでした。")

if __name__ == "__main__":
    convert_excel_to_markdown()
